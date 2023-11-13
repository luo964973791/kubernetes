# -*- coding: utf-8 -*-
#!/usr/bin/env python
#pip3 install openpyxl  && python3 执行.
import os
import socket
import openpyxl

def check_port(ip, port):
    if 0 <= port <= 65535:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(1)  # 设置超时时间，单位为秒

        try:
            sock.connect((ip, port))
            print(f"Port {port} is open on {ip}")
            return True
        except socket.error:
            print(f"Port {port} is not open on {ip}")
            return False
        finally:
            sock.close()
    else:
        print(f"Invalid port number: {port}")
        return False

# 清空 close.xlsx 文件
if os.path.exists('close.xlsx'):
    os.remove('close.xlsx')

# 创建一个新的工作表
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active
new_sheet.append(["ip", "Close Port"])

# 打开 Excel 文件
workbook = openpyxl.load_workbook('demo.xlsx')

# 遍历表格中的行
for row in workbook.active.iter_rows(min_row=2, values_only=True):  # 从第2行开始读取数据
    ip = row[0]
    ports = map(int, str(row[1]).split(',')) if ',' in str(row[1]) else [int(row[1])]

    # 检查端口是否开放
    closed_ports = [port for port in ports if not check_port(ip, port)]

    # 如果有不通的端口，写入新表格
    if closed_ports:
        closed_ports_str = ','.join(map(str, closed_ports))
        new_sheet.append([ip, closed_ports_str])

# 保存新表格
new_workbook.save('close.xlsx')
